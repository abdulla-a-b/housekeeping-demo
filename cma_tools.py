#!/usr/bin/env python3
"""
CLEAN MY AREA — Python companion
Good & Fast Packaging Co. Ltd.

Two jobs the browser should not do:

  seed    Generate the roster and zone CSVs to import into the Google Sheet
          (File > Import > Upload > Append to current sheet).

  report  Take the CSV the app exports and produce the formatted monthly
          Excel report for the Factory Manager.

Usage
    python cma_tools.py seed   --out ./seed
    python cma_tools.py report --csv clean-my-area-2026-08-31.csv --month 2026-08

Requires openpyxl for `report` only.
"""

import argparse
import csv
import os
from collections import defaultdict
from datetime import date, timedelta

CYCLE0 = date(2026, 8, 9)
START, END = date(2026, 8, 6), date(2026, 12, 31)
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

ROT = {
    (1, 6): ("Toilets & washrooms — Production Building (deep)", "Z-01"),
    (1, 1): ("Elastic production floor (deep)", "Z-12"),
    (1, 2): ("Canteen, dining hall & kitchen (deep)", "Z-03"),
    (1, 3): ("Yarn dyeing floor (deep)", "Z-13"),
    (2, 6): ("Toilets & washrooms — YD / Elastic / Utility (deep)", "Z-02"),
    (2, 1): ("Production Building GF + 1F (deep)", "Z-20"),
    (2, 2): ("Water points, medical room, child care (deep)", "Z-05"),
    (2, 3): ("ETP area & sludge bed to workshop (deep)", "Z-14"),
    (3, 6): ("Toilets & washrooms — Production Building (deep)", "Z-01"),
    (3, 1): ("Production Building 2F + 3F (deep)", "Z-21"),
    (3, 2): ("Central waste yard & segregation (deep)", "Z-11"),
    (3, 3): ("Fire exits, stairwells & escape routes (deep)", "Z-17"),
    (4, 6): ("Toilets & washrooms — YD / Elastic / Utility (deep)", "Z-02"),
    (4, 1): ("Warehouse — RM & FG (deep)", "Z-23"),
    (4, 2): ("Chemical & dyes store + spill station (deep)", "Z-16"),
    (4, 3): ("Machine underside & under-table (deep)", "Z-19"),
    (5, 6): ("Toilets & washrooms — Production Building (deep)", "Z-01"),
    (5, 1): ("Elastic production floor (deep)", "Z-12"),
    (5, 2): ("Changing, prayer, rest rooms + gate corridor", "Z-27"),
    (5, 3): ("Utility building — boiler, generator (deep)", "Z-15"),
    (6, 6): ("Toilets & washrooms — YD / Elastic / Utility (deep)", "Z-02"),
    (6, 1): ("PB + UTB roof, standing-water check", "Z-22"),
    (6, 2): ("Fans, light fittings, dust extraction + admin", "Z-18"),
    (6, 3): ("Yarn dyeing floor (deep)", "Z-13"),
}
FRI_DRY = {
    1: ("Internal roads & yard — deep scrub", "Z-25"),
    2: ("Warehouse high-level & racking clean", "Z-23"),
    3: ("Parking, perimeter & boundary strip", "Z-29"),
    4: ("Surface drains — annual de-silt", "Z-09"),
    5: ("Overhead tanks & reservoir — lid and base check", "Z-08"),
    6: ("Reserve / catch-up + audit-readiness walk", ""),
}
# id, name, risk, equipment colour, routine, area owner
ZONES = [
    ("Z-01", "Toilets & washrooms — Production Building", "CRITICAL", "RED", "3 rounds daily", "Admin / Housekeeping"),
    ("Z-02", "Toilets & washrooms — YD / Elastic / Utility", "CRITICAL", "RED", "3 rounds daily", "Admin / Housekeeping"),
    ("Z-03", "Canteen & dining hall", "CRITICAL", "GREEN", "After each meal", "Canteen In-charge"),
    ("Z-04", "Canteen kitchen & utensil wash", "CRITICAL", "GREEN", "Daily after service", "Canteen In-charge"),
    ("Z-05", "Drinking water points & filters", "CRITICAL", "GREEN", "Daily wipe + weekly sanitise", "Admin / Housekeeping"),
    ("Z-06", "Medical / first-aid room", "CRITICAL", "BLUE", "Daily", "Medical Officer"),
    ("Z-07", "Child care room", "CRITICAL", "BLUE", "2 rounds daily", "Welfare Officer"),
    ("Z-08", "Overhead water tanks & reservoir", "CRITICAL", "GREEN", "Weekly visual + quarterly clean", "Utility / Maintenance"),
    ("Z-09", "Surface drains & storm-water channels", "CRITICAL", "BLUE", "Weekly, 2x in monsoon", "Utility / Maintenance"),
    ("Z-10", "Waste collection points — floor bins", "HIGH", "BLUE", "2 rounds daily", "Section Heads"),
    ("Z-11", "Central waste yard & segregation", "HIGH", "BLUE", "Daily", "Store / Admin"),
    ("Z-12", "Elastic production floor", "HIGH", "YELLOW", "Layer 1 + daily sweep", "Elastic Section Head"),
    ("Z-13", "Yarn dyeing floor", "HIGH", "YELLOW", "Layer 1 + daily", "Yarn Dyeing Head"),
    ("Z-14", "ETP area & sludge bed to workshop", "HIGH", "BLUE", "Daily walk-through", "ETP In-charge"),
    ("Z-15", "Utility building — boiler, generator", "HIGH", "YELLOW", "Daily", "Utility In-charge"),
    ("Z-16", "Chemical & dyes store + spill station", "HIGH", "YELLOW", "Weekly", "Store In-charge"),
    ("Z-17", "Fire exits, stairwells & escape routes", "HIGH", "BLUE", "Daily obstruction check", "HSE Officer"),
    ("Z-18", "Fans, light fittings, dust extraction", "HIGH", "YELLOW", "Monthly", "Maintenance"),
    ("Z-19", "Machine underside & under-table", "HIGH", "YELLOW", "Layer 1 weekly", "Section Heads"),
    ("Z-20", "Production Building — ground + 1st floor", "MEDIUM", "YELLOW", "Daily sweep", "Section Heads"),
    ("Z-21", "Production Building — 2nd + 3rd floor", "MEDIUM", "YELLOW", "Daily sweep", "Section Heads"),
    ("Z-22", "PB + UTB roof, standing-water check", "MEDIUM", "BLUE", "Weekly in monsoon", "Maintenance"),
    ("Z-23", "Warehouse — raw material & finished goods", "MEDIUM", "YELLOW", "Daily aisle sweep", "Store In-charge"),
    ("Z-24", "Security gate to canteen corridor", "MEDIUM", "BLUE", "Daily", "Admin"),
    ("Z-25", "Internal roads & yard", "MEDIUM", "BLUE", "Daily 07:00-10:00", "Admin"),
    ("Z-26", "Landscaping — grass & bush cutting", "LOW", "NONE", "Weekly Apr-Oct", "Admin"),
    ("Z-27", "Changing, prayer & rest rooms", "MEDIUM", "BLUE", "Daily", "Welfare Officer"),
    ("Z-28", "Admin office & meeting rooms", "LOW", "BLUE", "Daily", "Admin"),
    ("Z-29", "Parking, perimeter & boundary", "LOW", "NONE", "Weekly", "Security / Admin"),
]


def third_thursday(y, m):
    d, n = date(y, m, 1), 0
    while True:
        if d.weekday() == 3:
            n += 1
            if n == 3:
                return d
        d += timedelta(days=1)


def phase_of(d):
    for cutoff, tag in [(date(2026, 8, 8), "P0"), (date(2026, 9, 4), "P1"),
                        (date(2026, 10, 3), "P2"), (date(2026, 10, 31), "P3"),
                        (date(2026, 11, 30), "P4"), (date(2026, 12, 31), "P5")]:
        if d <= cutoff:
            return tag
    return "—"


def assignment(d):
    wd, cyc = d.weekday(), (((d - CYCLE0).days // 7) % 6 + 1 if d >= CYCLE0 else 0)
    monsoon = 4 <= d.month <= 10
    if wd == 5:
        return "WEEKLY HOLIDAY", "", cyc
    if d < CYCLE0:
        return "Phase 0 — mobilisation, no rotation yet", "", cyc
    if d == third_thursday(d.year, d.month):
        return "MONTHLY TRAINING — 45 minutes, after lunch", "TRN", cyc
    if wd == 0:
        if monsoon or cyc % 2 == 1:
            return "Landscaping — grass & bush cutting", "Z-26", cyc
        return "Reserve / catch-up + Layer 1 coaching walk", "", cyc
    if wd == 4:
        if monsoon:
            return "Drain, roof & vector control", "Z-09", cyc
        z = FRI_DRY[cyc]
        return z[0], z[1], cyc
    z = ROT.get((cyc, wd), ("", ""))
    return z[0], z[1], cyc


def cmd_seed(out):
    os.makedirs(out, exist_ok=True)

    p = os.path.join(out, "roster.csv")
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["date", "day", "phase", "cycle_week", "deep_clean_zone",
                    "zone_id", "lead", "verifier", "done", "score"])
        d, n = START, 0
        while d <= END:
            zone, zid, cyc = assignment(d)
            w.writerow([d.isoformat(), DAYS[d.weekday()], phase_of(d),
                        cyc or "", zone, zid, "", "", "", ""])
            d += timedelta(days=1)
            n += 1
    print(f"  roster.csv   {n} rows  ({START} to {END})")

    p = os.path.join(out, "zones.csv")
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["zone_id", "zone", "risk_class", "equipment_colour",
                    "routine_frequency", "area_owner"])
        w.writerows(ZONES)
    print(f"  zones.csv    {len(ZONES)} rows")

    crit = sum(1 for z in ZONES if z[2] == "CRITICAL")
    print(f"\nSeed written to {out}/")
    print(f"{crit} zones are CRITICAL risk — they run daily regardless of the rotation.")
    print("Import into the Google Sheet: File > Import > Upload > Append to current sheet.")


def cmd_report(csv_path, month):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise SystemExit("openpyxl is required for `report`:  pip install openpyxl")

    rows = list(csv.DictReader(open(csv_path, encoding="utf-8-sig")))
    ros = [r for r in rows if r["type"] == "roster" and r["date"].startswith(month)]
    ins = [r for r in rows if r["type"] == "inspection" and r["date"].startswith(month)]
    vec = [r for r in rows if r["type"] == "vector" and r["date"].startswith(month)]

    work = [r for r in ros if date.fromisoformat(r["date"]).weekday() != 5]
    done = sum(1 for r in work if r["done"] == "Y")
    adh = done / len(work) if work else None
    scores = [float(r["score"]) for r in ins if r["score"]]
    avg = sum(scores) / len(scores) if scores else None
    closed = sum(1 for r in vec if r["done"] == "Y")

    by_zone = defaultdict(list)
    for r in ins:
        if r["score"]:
            by_zone[r["zone"]].append(float(r["score"]))

    NAVY, GREEN, RED, AMBER = "101211", "12924A", "E23C1E", "FFC400"
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"CLEAN MY AREA — {month}"
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws["A2"] = "Good & Fast Packaging Co. Ltd. — generated from the app record"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    for col, wdt in zip("ABCDE", (34, 16, 16, 16, 30)):
        ws.column_dimensions[col].width = wdt

    r = 4
    for i, h in enumerate(("Indicator", "Value", "Target", "Status", "Note"), start=1):
        cell = ws.cell(r, i, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.border = box
        cell.alignment = Alignment(horizontal="center")

    def line(label, value, target, ok, note):
        nonlocal r
        r += 1
        for i, v in enumerate((label, value, target, "ON TARGET" if ok else "BELOW", note), start=1):
            cell = ws.cell(r, i, v)
            cell.font = Font(name="Arial", size=10, bold=(i == 1))
            cell.border = box
            cell.alignment = Alignment(horizontal="center" if 2 <= i <= 4 else "left", wrap_text=True)
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=GREEN if ok else RED)
        ws.cell(r, 4).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    line("Schedule adherence", f"{adh:.0%}" if adh is not None else "—", "95%",
         adh is not None and adh >= .95, f"{done} of {len(work)} working days signed off")
    line("Average inspection score", f"{avg:.1f}" if avg is not None else "—", "40 / 50",
         avg is not None and avg >= 40, f"{len(ins)} inspections recorded")
    line("Zones below 30", sum(1 for s in scores if s < 30), "0",
         not any(s < 30 for s in scores), "Two in a row triggers a written corrective plan")
    line("Water sites found", len(vec), "—", True, "Finding water is not the failure")
    line("Closed within 24 hours", f"{closed/len(vec):.0%}" if vec else "—", "100%",
         bool(vec) and closed == len(vec), f"{closed} of {len(vec)} sites")

    r += 3
    ws.cell(r, 1, "ZONE SCORES").font = Font(name="Arial", size=12, bold=True, color=NAVY)
    r += 1
    for i, h in enumerate(("Zone", "Inspections", "Average", "Lowest", "Rating"), start=1):
        cell = ws.cell(r, i, h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.border = box
        cell.alignment = Alignment(horizontal="center")
    zlook = {z[0]: z[1] for z in ZONES}
    for zid in sorted(by_zone):
        vals = by_zone[zid]
        a = sum(vals) / len(vals)
        r += 1
        rating = "EXCELLENT" if a >= 45 else "PASS" if a >= 40 else "IMPROVE" if a >= 30 else "FAIL"
        for i, v in enumerate((f"{zid} · {zlook.get(zid, '')}", len(vals),
                               round(a, 1), min(vals), rating), start=1):
            cell = ws.cell(r, i, v)
            cell.font = Font(name="Arial", size=10)
            cell.border = box
            cell.alignment = Alignment(horizontal="center" if i > 1 else "left")
        ws.cell(r, 5).fill = PatternFill(
            "solid", fgColor=GREEN if a >= 40 else AMBER if a >= 30 else RED)

    out = f"Clean_My_Area_Report_{month}.xlsx"
    wb.save(out)
    print(f"Written: {out}")
    print(f"  {len(work)} working days, {done} signed off"
          + (f" ({adh:.0%})" if adh is not None else ""))
    print(f"  {len(ins)} inspections" + (f", average {avg:.1f}/50" if avg is not None else ""))
    print(f"  {len(vec)} water sites, {closed} closed within 24 hours")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Clean My Area — seed and report tools")
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("seed", help="generate roster.csv and zones.csv for the Google Sheet")
    s.add_argument("--out", default="./seed")
    rp = sub.add_parser("report", help="build the monthly Excel report from an app CSV export")
    rp.add_argument("--csv", required=True)
    rp.add_argument("--month", required=True, help="YYYY-MM")
    a = ap.parse_args()
    if a.cmd == "seed":
        cmd_seed(a.out)
    else:
        cmd_report(a.csv, a.month)
